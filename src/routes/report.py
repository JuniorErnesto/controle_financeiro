"""
src/routes/report.py

Este módulo contém as rotas para geração de relatórios financeiros em PDF e Excel.
Ele utiliza os modelos SQLAlchemy para interagir com o banco de dados MySQL.
"""

from flask import Blueprint, request, jsonify, make_response
from flask_login import login_required, current_user
from sqlalchemy import func, and_
from src.models.user import db, User, Transaction  # Alterado para modelos SQLAlchemy
from datetime import datetime, timedelta, time
from fpdf import FPDF
import io
import pandas as pd

report_bp = Blueprint("report", __name__)

class PDF(FPDF):
    """Classe personalizada para gerar o cabeçalho e rodapé do PDF."""
    def header(self):
        self.set_font("Arial", "B", 12)
        self.cell(0, 10, "Relatório Financeiro", 0, 1, "C")
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Página {self.page_no()}/{{nb}}", 0, 0, "C")

    def chapter_title(self, title):
        self.set_font("Arial", "B", 12)
        self.cell(0, 10, title, 0, 1, "L")
        self.ln(5)

    def chapter_body(self, body):
        self.set_font("Arial", "", 10)
        self.multi_cell(0, 5, body)
        self.ln()

    def add_table(self, header, data):
        self.set_fill_color(200, 220, 255)
        self.set_text_color(0,0,0)
        self.set_draw_color(128,128,128)
        self.set_line_width(0.3)
        self.set_font("Arial", "B", 10)
        
        col_widths = [30, 70, 30, 25, 30] # Ajustado para corresponder aos dados

        for i, header_text in enumerate(header):
            self.cell(col_widths[i], 7, header_text, 1, 0, "C", 1)
        self.ln()

        self.set_font("Arial", "", 9)
        fill = False
        for row in data:
            self.set_fill_color(240, 240, 240) if fill else self.set_fill_color(255, 255, 255)
            self.cell(col_widths[0], 6, str(row[0]), "LRB", 0, "L", fill)
            self.cell(col_widths[1], 6, str(row[1]), "LRB", 0, "L", fill)
            self.cell(col_widths[2], 6, str(row[2]), "LRB", 0, "L", fill)
            self.cell(col_widths[3], 6, str(row[3]), "LRB", 0, "C", fill)
            self.cell(col_widths[4], 6, f"{row[4]:.2f}", "LRB", 0, "R", fill)
            self.ln()
            fill = not fill
        self.cell(sum(col_widths), 0, "", "T")
        self.ln()

def _get_filtered_transactions_and_summary(user_id):
    """Busca transações filtradas e calcula o resumo financeiro usando SQLAlchemy."""
    query = db.session.query(Transaction).filter_by(user_id=user_id)
    
    period = request.args.get("period")
    start_date_str = request.args.get("start_date")
    end_date_str = request.args.get("end_date")

    filter_description = "Período: Todas as transações"
    
    filters = []

    if period:
        today_dt = datetime.utcnow()
        today_date = today_dt.date()
        if period == "day":
            start_dt = datetime.combine(today_date, time.min)
            end_dt = datetime.combine(today_date, time.max)
            filters.append(Transaction.date >= start_dt)
            filters.append(Transaction.date <= end_dt)
            filter_description = f"Período: Dia ({today_date.strftime('%d/%m/%Y')})"
        elif period == "week":
            start_of_week_date = today_date - timedelta(days=today_date.weekday())
            end_of_week_date = start_of_week_date + timedelta(days=6)
            start_dt = datetime.combine(start_of_week_date, time.min)
            end_dt = datetime.combine(end_of_week_date, time.max)
            filters.append(Transaction.date >= start_dt)
            filters.append(Transaction.date <= end_dt)
            filter_description = f"Período: Semana ({start_of_week_date.strftime('%d/%m/%Y')} - {end_of_week_date.strftime('%d/%m/%Y')})"
        elif period == "month":
            start_of_month_date = today_date.replace(day=1)
            next_month_date = (start_of_month_date + timedelta(days=32)).replace(day=1)
            end_of_month_date = next_month_date - timedelta(days=1)
            start_dt = datetime.combine(start_of_month_date, time.min)
            end_dt = datetime.combine(end_of_month_date, time.max)
            filters.append(Transaction.date >= start_dt)
            filters.append(Transaction.date <= end_dt)
            filter_description = f"Período: Mês ({start_of_month_date.strftime('%B/%Y')})"
        elif period == "year":
            start_of_year_date = today_date.replace(month=1, day=1)
            end_of_year_date = today_date.replace(month=12, day=31)
            start_dt = datetime.combine(start_of_year_date, time.min)
            end_dt = datetime.combine(end_of_year_date, time.max)
            filters.append(Transaction.date >= start_dt)
            filters.append(Transaction.date <= end_dt)
            filter_description = f"Período: Ano ({start_of_year_date.year})"
    elif start_date_str and end_date_str:
        try:
            start_date_dt = datetime.fromisoformat(start_date_str.replace("Z", "+00:00"))
            end_date_dt = datetime.fromisoformat(end_date_str.replace("Z", "+00:00"))
            
            start_dt = datetime.combine(start_date_dt.date(), time.min)
            end_dt = datetime.combine(end_date_dt.date(), time.max)
            
            filters.append(Transaction.date >= start_dt)
            filters.append(Transaction.date <= end_dt)
            filter_description = f"Período: {start_dt.strftime('%d/%m/%Y')} a {end_dt.strftime('%d/%m/%Y')}"
        except ValueError:
            raise ValueError("Formato de data inválido para start_date ou end_date. Use o formato ISO.")

    if filters:
        query = query.filter(and_(*filters))

    transactions_list = query.order_by(Transaction.date.asc()).all()
    
    # Cálculos de totais com SQLAlchemy
    total_receitas_query = db.session.query(func.sum(Transaction.amount)).filter_by(user_id=user_id, type="receita")
    total_despesas_query = db.session.query(func.sum(Transaction.amount)).filter_by(user_id=user_id, type="despesa")

    if filters: # Aplicar os mesmos filtros de data para os totais
        total_receitas_query = total_receitas_query.filter(and_(*filters))
        total_despesas_query = total_despesas_query.filter(and_(*filters))

    total_receitas = total_receitas_query.scalar() or 0.0
    total_despesas = total_despesas_query.scalar() or 0.0
    saldo = total_receitas - total_despesas

    return transactions_list, total_receitas, total_despesas, saldo, filter_description

@report_bp.route("/report/pdf", methods=["GET"])
@login_required
def generate_pdf_report():
    try:
        transactions, total_receitas, total_despesas, saldo, filter_description = _get_filtered_transactions_and_summary(current_user.id)
    except ValueError as e:
        return jsonify({"message": str(e)}), 400
    except Exception as e:
        print(f"Erro ao gerar relatório PDF: {e}") # Log do erro
        return jsonify({"message": "Erro interno ao gerar relatório PDF."}), 500

    pdf = PDF()
    pdf.alias_nb_pages()
    pdf.add_page()
    pdf.set_font("Arial", "", 10)

    pdf.chapter_title("Resumo Financeiro")
    pdf.chapter_body(f"Usuário: {current_user.username}")
    pdf.chapter_body(filter_description)
    pdf.ln(5)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 7, "Total de Receitas:", 0, 0)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 7, f"R$ {total_receitas:.2f}", 0, 1)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 7, "Total de Despesas:", 0, 0)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 7, f"R$ {total_despesas:.2f}", 0, 1)
    pdf.set_font("Arial", "B", 10)
    pdf.cell(40, 7, "Saldo Final:", 0, 0)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 7, f"R$ {saldo:.2f}", 0, 1)
    pdf.ln(10)

    pdf.chapter_title("Detalhes das Transações")
    table_header = ["Data", "Descrição", "Categoria", "Tipo", "Valor (R$)"]
    table_data = []
    for t in transactions:
        table_data.append([
            t.date.strftime("%d/%m/%Y"),
            t.description,
            t.category if t.category else "-",
            t.type.capitalize(),
            t.amount
        ])
    
    if table_data:
        pdf.add_table(table_header, table_data)
    else:
        pdf.chapter_body("Nenhuma transação encontrada para o período selecionado.")

    pdf_output_bytes = pdf.output(dest='S') # Retorna bytes diretamente com fpdf2
    pdf_output = io.BytesIO(pdf_output_bytes)

    response = make_response(pdf_output.read())
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = "attachment; filename=relatorio_financeiro.pdf"
    return response

@report_bp.route("/report/excel", methods=["GET"])
@login_required
def generate_excel_report():
    try:
        transactions, total_receitas, total_despesas, saldo, filter_description = _get_filtered_transactions_and_summary(current_user.id)
    except ValueError as e:
        return jsonify({"message": str(e)}), 400
    except Exception as e:
        print(f"Erro ao gerar relatório Excel: {e}") # Log do erro
        return jsonify({"message": "Erro interno ao gerar relatório Excel."}), 500

    data_for_df = {
        "Data": [t.date.strftime("%d/%m/%Y") for t in transactions],
        "Descrição": [t.description for t in transactions],
        "Categoria": [t.category if t.category else "-" for t in transactions],
        "Tipo": [t.type.capitalize() for t in transactions],
        "Valor (R$)": [t.amount for t in transactions]
    }
    df_transactions = pd.DataFrame(data_for_df)

    summary_data = {
        "Descrição": ["Total de Receitas", "Total de Despesas", "Saldo Final"],
        "Valor (R$)": [total_receitas, total_despesas, saldo]
    }
    df_summary = pd.DataFrame(summary_data)

    excel_output = io.BytesIO()
    with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
        df_summary.to_excel(writer, sheet_name="Resumo Financeiro", index=False, startrow=3)
        sheet_summary = writer.sheets["Resumo Financeiro"]
        
        sheet_summary["A1"] = "Resumo Financeiro"
        # Corrigir acesso à fonte para openpyxl
        try:
            from openpyxl.styles import Font
            sheet_summary["A1"].font = Font(bold=True, size=14)
        except ImportError:
            print("openpyxl.styles.Font não encontrado, pulando formatação de fonte.")
            pass # Continuar sem formatação se houver problema com a importação

        sheet_summary["A2"] = f"Usuário: {current_user.username}"
        sheet_summary["A3"] = filter_description

        df_transactions.to_excel(writer, sheet_name="Detalhes das Transações", index=False)
        
        # Remover a planilha "Sheet1" padrão se ela foi criada e estiver vazia
        if "Sheet1" in writer.book.sheetnames:
            std_sheet = writer.book["Sheet1"]
            if std_sheet.max_row == 1 and std_sheet.max_column == 1 and std_sheet["A1"].value is None:
                 writer.book.remove(std_sheet)
            elif writer.book.active == std_sheet and len(writer.book.sheetnames) > 1:
                active_sheet_name = "Resumo Financeiro"
                if active_sheet_name in writer.book.sheetnames:
                    writer.book.active = writer.book.sheetnames.index(active_sheet_name)
                if "Sheet1" in writer.book.sheetnames: # Checa novamente
                     writer.book.remove(writer.book["Sheet1"])

    excel_output.seek(0)

    response = make_response(excel_output.read())
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Content-Disposition"] = "attachment; filename=relatorio_financeiro.xlsx"
    return response

